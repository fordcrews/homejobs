#!/usr/bin/env python3
"""Build price-compare.csv and price-compare.xlsx from prices.json."""
from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent

NAMES = {
    "fencePrices.v1:p4": ("Wood fence", "4x4x8 treated post (ground contact)", "ea"),
    "fencePrices.v1:p6": ("Wood fence", "6x6x8 treated post (corner/end/gate)", "ea"),
    "fencePrices.v1:rail": ("Wood fence", "2x4x8 treated rail", "ea"),
    "fencePrices.v1:board": ("Wood fence", "6 ft 1x6 treated dog-ear picket", "ea"),
    "fencePrices.v1:conc": ("Wood fence", "Concrete mix 50-80 lb bag", "bag"),
    "fencePrices.v1:cap": ("Wood fence", "Post cap (pyramid)", "ea"),
    "fencePrices.v1:hinge": ("Wood fence", "Heavy T-hinge pair", "pair"),
    "fencePrices.v1:latch": ("Wood fence", "Gate latch", "ea"),
    "fencePrices.v1:drop": ("Wood fence", "Cane bolt / drop rod", "ea"),
    "fencePrices.v1:screws": ("Wood fence", "Exterior screws #8 x 2 in (1 lb box)", "box"),
    "chainlinkPrices.v1:term": ("Cyclone fence", "Terminal post 2-3/8 in OD", "ea"),
    "chainlinkPrices.v1:line": ("Cyclone fence", "Line post 1-5/8 in OD", "ea"),
    "chainlinkPrices.v1:rail": ("Cyclone fence", "Top rail 1-3/8 in x 10.5 ft", "ea"),
    "chainlinkPrices.v1:mesh": ("Cyclone fence", "4 ft x 50 ft 11.5 ga chain-link fabric", "roll"),
    "chainlinkPrices.v1:bar": ("Cyclone fence", "Tension bar", "ea"),
    "chainlinkPrices.v1:band": ("Cyclone fence", "Tension band", "ea"),
    "chainlinkPrices.v1:brace": ("Cyclone fence", "Brace band", "ea"),
    "chainlinkPrices.v1:capl": ("Cyclone fence", "Loop cap (line post)", "ea"),
    "chainlinkPrices.v1:capt": ("Cyclone fence", "Dome cap (terminal)", "ea"),
    "chainlinkPrices.v1:rend": ("Cyclone fence", "Rail end", "ea"),
    "chainlinkPrices.v1:tie": ("Cyclone fence", "Aluminum tie wire (bag ~100)", "bag"),
    "chainlinkPrices.v1:wire": ("Cyclone fence", "Bottom tension wire", "lf"),
    "chainlinkPrices.v1:conc": ("Cyclone fence", "Concrete mix 50-80 lb bag", "bag"),
    "chainlinkPrices.v1:wgate": ("Cyclone fence", "Walk gate (framed)", "ea"),
    "chainlinkPrices.v1:dgate": ("Cyclone fence", "Double drive gate", "ea"),
    "floorPrices.v1:flr": ("Floors", "LVP / laminate / hardwood carton (~20 sf)", "carton"),
    "floorPrices.v1:pad": ("Floors", "Underlayment or carpet pad", "roll"),
    "floorPrices.v1:trans": ("Floors", "Transition / T-mold / threshold", "ea"),
    "floorPrices.v1:qtr": ("Floors", "Quarter-round / shoe", "lf"),
    "floorPrices.v1:glue": ("Floors", "Wood adhesive / nails", "tube"),
    "floorPrices.v1:tile": ("Floors", "Ceramic / porcelain tile carton", "carton"),
    "floorPrices.v1:thin": ("Floors", "Thinset / mortar 50 lb", "bag"),
    "floorPrices.v1:grout": ("Floors", "Sanded grout 25 lb", "bag"),
    "floorPrices.v1:spac": ("Floors", "Tile spacers", "bag"),
    "floorPrices.v1:uw": ("Floors", "Cement board 3x5 sheet", "sheet"),
    "floorPrices.v1:base": ("Floors", "Tile base / shoe", "lf"),
    "floorPrices.v1:carp": ("Floors", "Carpet 12 ft goods", "lf"),
    "floorPrices.v1:tack": ("Floors", "Tack strip", "lf"),
    "paintPrices.v1:wall": ("Paint", "Wall paint eggshell gallon", "gal"),
    "paintPrices.v1:ceil": ("Paint", "Ceiling paint gallon", "gal"),
    "paintPrices.v1:prim": ("Paint", "Primer gallon", "gal"),
    "paintPrices.v1:trim": ("Paint", "Trim enamel gallon", "gal"),
    "paintPrices.v1:tape": ("Paint", "Painter tape 1.88 in", "roll"),
    "paintPrices.v1:plastic": ("Paint", "Drop cloth / plastic", "ea"),
    "paintPrices.v1:roll": ("Paint", "9 in roller cover 3-pack", "pk"),
    "paintPrices.v1:brush": ("Paint", "2-2.5 in sash brush", "ea"),
    "paintPrices.v1:tray": ("Paint", "Tray + liner", "set"),
}


def num(v):
    if v in ("", None):
        return None
    try:
        return float(v)
    except ValueError:
        return None


def main() -> None:
    prices = json.loads((HERE / "prices.json").read_text(encoding="utf-8"))
    catalog = json.loads((HERE / "catalog.json").read_text(encoding="utf-8"))
    url_by_key = {}
    for it in catalog.get("items", []):
        for k in it.get("keys", []):
            url_by_key.setdefault(k, {"hd": it.get("hd_url", ""), "lw": it.get("lw_url", "")})

    rows = []
    for key, (job, name, unit) in NAMES.items():
        table, iid = key.split(":", 1)
        p = prices.get(table, {}).get(iid, {})
        urls = url_by_key.get(key, {})
        rows.append(
            {
                "job": job,
                "item": name,
                "unit": unit,
                "key": key,
                "home_depot": p.get("hd", ""),
                "lowes": p.get("lw", ""),
                "local1": p.get("l1", p.get("ml", "")),
                "local2": p.get("l2", ""),
                "hd_url": urls.get("hd", ""),
                "lowes_url": urls.get("lw", ""),
            }
        )

    csv_path = HERE / "price-compare.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "job",
                "item",
                "unit",
                "key",
                "home_depot",
                "lowes",
                "local1",
                "local2",
                "hd_url",
                "lowes_url",
            ],
        )
        w.writeheader()
        w.writerows(rows)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Stores"
    ws0["A1"] = "Homejobs price comparison"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A2"] = "HD / Lowes catalog as of"
    ws0["B2"] = prices.get("updated", "")
    ws0["A3"] = "Area"
    ws0["B3"] = "Jackson, MS"
    ws0["A5"] = "Yellow cells are yours — type local shop names and prices."
    ws0["A5"].font = Font(italic=True, color="806000")
    for col, h in enumerate(("Store", "Name (edit)", "Phone / notes"), 1):
        c = ws0.cell(7, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="243447")
    ws0["A8"] = "Home Depot"
    ws0["B8"] = "Home Depot (Jackson)"
    ws0["A9"] = "Lowes"
    ws0["B9"] = "Lowes (Jackson)"
    ws0["A10"] = "Local 1"
    ws0["B10"] = "Local lumber / fence shop 1"
    ws0["C10"] = ""
    ws0["A11"] = "Local 2"
    ws0["B11"] = "Local lumber / fence shop 2"
    yellow = PatternFill("solid", fgColor="FFF2CC")
    for addr in ("B10", "B11", "C10", "C11"):
        ws0[addr].fill = yellow
    ws0.column_dimensions["A"].width = 16
    ws0.column_dimensions["B"].width = 36
    ws0.column_dimensions["C"].width = 40

    ws = wb.create_sheet("Prices", 0)
    headers = [
        "Job",
        "Item",
        "Unit",
        "Key",
        "Home Depot",
        "Lowes",
        "Local 1",
        "Local 2",
        "Cheapest",
        "Cheapest store",
        "HD URL",
        "Lowes URL",
    ]
    thin = Border(
        left=Side(style="thin", color="D5DDE4"),
        right=Side(style="thin", color="D5DDE4"),
        top=Side(style="thin", color="D5DDE4"),
        bottom=Side(style="thin", color="D5DDE4"),
    )
    hdr = PatternFill("solid", fgColor="243447")
    hd_fill = PatternFill("solid", fgColor="FDE8D8")
    lw_fill = PatternFill("solid", fgColor="D6EAF8")
    l_fill = PatternFill("solid", fgColor="FFF2CC")
    win_fill = PatternFill("solid", fgColor="C6EFCE")
    ws.append(headers)
    for col in range(1, 13):
        c = ws.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    for i, r in enumerate(rows, 2):
        ws.cell(i, 1, r["job"])
        ws.cell(i, 2, r["item"])
        ws.cell(i, 3, r["unit"])
        ws.cell(i, 4, r["key"])
        for col, key in ((5, "home_depot"), (6, "lowes"), (7, "local1"), (8, "local2")):
            cell = ws.cell(i, col, num(r[key]))
            cell.number_format = '"$"#,##0.00'
        ws.cell(i, 5).fill = hd_fill
        ws.cell(i, 6).fill = lw_fill
        ws.cell(i, 7).fill = l_fill
        ws.cell(i, 8).fill = l_fill
        cheap = ws.cell(i, 9)
        cheap.value = f"=IF(COUNT(E{i}:H{i})=0,\"\",MIN(E{i}:H{i}))"
        cheap.number_format = '"$"#,##0.00'
        cheap.fill = win_fill
        ws.cell(i, 10).value = (
            f'=IF(I{i}="","",INDEX($E$1:$H$1,MATCH(I{i},E{i}:H{i},0)))'
        )
        ws.cell(i, 11, r["hd_url"])
        ws.cell(i, 12, r["lowes_url"])
        for col in range(1, 13):
            ws.cell(i, col).border = thin
            ws.cell(i, col).alignment = Alignment(
                vertical="center", wrap_text=(col in (2, 11, 12))
            )

    last = 1 + len(rows)
    ws.auto_filter.ref = f"A1:L{last}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for i, w in enumerate([16, 48, 8, 28, 14, 12, 12, 12, 12, 16, 42, 42], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    note = wb.create_sheet("How to use")
    note["A1"] = "How to use"
    note["A1"].font = Font(bold=True, size=14)
    note["A3"] = "1. Stores tab: type real names for Local 1 and Local 2."
    note["A4"] = "2. Prices tab: yellow Local 1 / Local 2 are blank — fill from a mill or shop quote."
    note["A5"] = "3. Home Depot and Lowes are national web catalog prices (Jackson shelf may differ)."
    note["A6"] = "4. Cheapest and Cheapest store update when you type a number."
    note["A7"] = "5. Do not change the Key column — the website calculators use it."
    note["A8"] = "Rebuild this file after editing prices.json:  python build_compare_sheet.py"
    note.column_dimensions["A"].width = 110

    xlsx = HERE / "price-compare.xlsx"
    wb.save(xlsx)
    print(f"Wrote {len(rows)} rows")
    print(xlsx)
    print(csv_path)


if __name__ == "__main__":
    main()
